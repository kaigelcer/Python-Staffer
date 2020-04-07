import random
import datetime
from openpyxl import load_workbook


Entry = load_workbook('BookForProgramDirector.xlsx')
Data = load_workbook('BookForPython.xlsx')

DUTY_SCHEDULE_SHEET = Entry["Duty Schedule"]
STAFFING_SHEET = Entry["Official Staffing"]
STAFF_ROSTER_SHEET = Entry["Staff Roster"]

# Activity sheet constants
ACTIVITY_SHEET = Data["Activity Numbers"]
ACTIVITY_COLUMN = 1
MIN_STAFF_COLUMN = 2
NUM_STAFF_COLUMN = 3
MAX_STAFF_COLUMN = 4
ROPES_COLUMN = 5
WATERFRONT_COLUMN = 6
MIN_NLS_COLUMN = 7
APTLY_STAFFED_COLUMN = 8

# Compute number of activities
number_of_activities = 0
while (ACTIVITY_SHEET.cell(row=number_of_activities+2, column=1)).value:
    number_of_activities += 1

# Compute number of staff members
number_of_staff_members = 0
while (STAFF_ROSTER_SHEET.cell(row=number_of_staff_members+2, column=1)).value:
    number_of_staff_members += 1

time_of_day = ["Morning", "Afternoon"]

# For both morning and afternoon
for time in time_of_day:

    # If it's the morning, clear the previous days staffing
    if time == "Morning":

        for staff in range(2, number_of_staff_members + 2):
            for period in range(2, 7):
                STAFFING_SHEET.cell(row=staff, column=period, value="---")

    # Creates a list with a numerical value for each staff member, then shuffle it
    staff_list = []
    for number in range(2, number_of_staff_members + 2):
        staff_list.append(number)

    random.shuffle(staff_list)
    print(staff_list)

    # Set current number of staff at every activity to 0
    for activity in range(2, number_of_activities + 2):
        ACTIVITY_SHEET.cell(row=activity, column=NUM_STAFF_COLUMN, value=0)



Entry.save("BookForProgramDirector.xlsx")
Data.save("BookForPython.xlsx")
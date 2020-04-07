import datetime
import openpyxl
from openpyxl import styles
from openpyxl.utils import get_column_letter
import random

# Open Workbooks
Entry = openpyxl.load_workbook('BookForProgramDirector.xlsx')
Data = openpyxl.load_workbook('BookForPython.xlsx')

# Constants for Session Dates sheet
SESSION_DATES_SHEET = Entry["Session Dates"]
SESSION_COLUMN = 1
SESSION_START_COLUMN = 2
SESSION_END_COLUMN = 3
A1_ROW = 2
A2_ROW = 3
B1_ROW = 4
B2_ROW = 5
C_ROW = 6
A1_START = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_START_COLUMN)
A2_START = SESSION_DATES_SHEET.cell(row=A2_ROW, column=SESSION_START_COLUMN)
A1_END = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_END_COLUMN)


# Constants for Full Day, Half Day, and Duty sheets
FULL_DAY_SHEET = Entry["Full Day"]
HALF_DAY_SHEET = Entry["Half Day"]
DUTY_SCHEDULE_SHEET = Entry["Duty Schedule"]
DATE_COLUMN = 2
DATE_ROW = 2
NAME_COLUMN_A = 2
TYPE_COLUMN_A = 1

# Constants for Staff Roster
STAFF_ROSTER_SHEET = Entry["Staff Roster"]
NAME_COLUMN = 2
TYPE_COLUMN = 3
totalStaff = (STAFF_ROSTER_SHEET.cell(row=2, column=12)).value
typeList = ["JB", "JG", "BB", "BG", "IB", "IG", "SB", "SG", "SSB", "SSG", "LT", "SPEC", "TRIPPER"]


# Constants for Jinci Boys and Jinci Girls Rotation, Jinci Boys and Jinci Girls Schedule
JB_ROTATION_SHEET = Data["Jinci Boys Rotation"]
JG_ROTATION_SHEET = Data["Jinci Girls Rotation"]
JB_SCHEDULE_SHEET = Entry["Jinci Boys Schedule"]
JG_SCHEDULE_SHEET = Entry["Jinci Girls Schedule"]
ROTATION_COLUMN = 1
CABIN_COLUMN = 2
OFFSET_COLUMN = 3
CABIN1_ROW = 2
CABIN2_ROW = 3
CABIN3_ROW = 4
TotalCabinsJB = (JB_ROTATION_SHEET.cell(row=2, column=4)).value


def date_mon_dd(date_to_print):
    return date_to_print.strftime("%b-%d")


def date_plus_1(date):
    in_date_form = datetime.datetime.strptime(date, "%b-%d")
    plus_1_day = datetime.timedelta(days=1)
    new_date = in_date_form + plus_1_day
    return new_date.strftime("%b-%d")


def date_minus_1(date):
    in_date_form = datetime.datetime.strptime(date, "%b-%d")
    minus_1_day = datetime.timedelta(days=-1)
    new_date = in_date_form + minus_1_day
    return new_date.strftime("%b-%d")


def print_dates_to_column_from_table(start_row, end_row, start_column, end_column, read_sheet, write_column,
                                     write_start, write_sheet):
    for i in range(start_row, end_row + 1):
        session = (read_sheet.cell(row=i, column=start_column - 1)).value
        write_sheet.cell(row=write_start, column=write_column - 1, value=session)
        start_date = (read_sheet.cell(row=i, column=start_column)).value
        end_date = (read_sheet.cell(row=i, column=end_column)).value
        date_1 = datetime.datetime.strptime(start_date, "%b-%d")
        date_2 = datetime.datetime.strptime(end_date, "%b-%d")
        delta_t = date_2 - date_1
        total_days = int((delta_t.total_seconds()) / 86400)
        for j in range(0, total_days):
            num_days = datetime.timedelta(days=j)
            date_to_print = date_1 + num_days
            write_sheet.cell(row=write_start, column=write_column, value=date_mon_dd(date_to_print))
            write_start += 1


def print_dates_to_row_from_table(start_row, end_row, start_column, end_column, read_sheet, write_row, write_start,
                                  write_sheet):
    for i in range(start_row, end_row + 1):
        session = (read_sheet.cell(row=i, column=start_column - 1)).value
        write_sheet.cell(row=write_row - 1, column=write_start, value=session)
        start_date = (read_sheet.cell(row=i, column=start_column)).value
        end_date = (read_sheet.cell(row=i, column=end_column)).value
        date_1 = datetime.datetime.strptime(start_date, "%b-%d")
        date_2 = datetime.datetime.strptime(end_date, "%b-%d")
        delta_t = date_2 - date_1
        total_days = int((delta_t.total_seconds()) / 86400)
        for j in range(0, total_days):
            num_days = datetime.timedelta(days=j)
            date_to_print = date_1 + num_days
            write_sheet.cell(row=write_row, column=write_start, value=date_mon_dd(date_to_print))
            write_sheet.cell(row=write_row, column=write_start).alignment = styles.Alignment(text_rotation=90,
                                                                                             horizontal="center",
                                                                                             shrinkToFit=True)
            cell = write_sheet.cell(row=write_row, column=write_start)
            write_sheet.column_dimensions[get_column_letter(cell.column)].width = 4
            write_start += 1


def sort_names_by_type(read_sheet, read_start, read_end, read_column, write_sheet, write_start, write_column,
                       type_list, type_column_read, type_column_write):
    for j in range(0, len(type_list)):
        write_sheet.cell(row=write_start, column=type_column_write, value=type_list[j])
        for i in range(read_start, read_end + 1):
            if (read_sheet.cell(row=i, column=type_column_read)).value == type_list[j]:
                staff_name = (read_sheet.cell(row=i, column=read_column)).value
                write_sheet.cell(row=write_start, column=write_column, value=staff_name)
                write_start += 1


def create_list_from_column(column_start, read_sheet, read_column):
    my_list = []
    acc = column_start
    while (read_sheet.cell(row=acc, column=read_column)).value:
        my_list.append((read_sheet.cell(row=acc, column=read_column)).value)
        acc += 1
    return my_list


def create_jinci_schedules(start_column, read_sheet, write_sheet, entry_list):
    for i in range(2, TotalCabinsJB + 2):
        cabin = (read_sheet.cell(row=i, column=CABIN_COLUMN)).value
        offset = (read_sheet.cell(row=i, column=OFFSET_COLUMN)).value
        print_dates_to_column_from_table(A1_ROW, C_ROW, SESSION_START_COLUMN, SESSION_END_COLUMN, SESSION_DATES_SHEET,
                                         start_column, 3, write_sheet)
        start_column += 1
        write_sheet.cell(row=1, column=start_column, value=cabin)
        for j in range(1, 6):
            write_sheet.cell(row=2, column=start_column, value=("Period " + str(j)))
            start_column += 1
        start_column -= 5
        row = 3
        acc = 1
        while (write_sheet.cell(row=row, column=2)).value:
            for k in range(1, 6):
                if (write_sheet.cell(row=2, column=start_column)).value == "Period 3":
                    write_sheet.cell(row=row, column=start_column, value="SWIM")
                else:
                    rotation = ((acc + offset) % len(entry_list))
                    write_sheet.cell(row=row, column=start_column, value=(entry_list[rotation - 1]))
                    acc += 1
                start_column += 1
            row += 1
            start_column -= 5
        start_column += 7


def sum_column_while_value(sheet, start_row, sum_column, value_column):
    total = 0
    while (sheet.cell(row=start_row, column=value_column)).value:
        total += (sheet.cell(row=start_row, column=sum_column)).value
        start_row += 1
    return total


def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3


def add_names_to_lists_if(list1, list2, boolean_column1, boolean_column2, name, name_column, write_sheet):
    start_row = 2
    while (write_sheet.cell(row=start_row, column=name_column)).value:
        if (write_sheet.cell(row=start_row, column=name_column)).value == name:
            if (write_sheet.cell(row=start_row, column=boolean_column1)).value == "Y":
                list1.append(name)
            if (write_sheet.cell(row=start_row, column=boolean_column2)).value == "Y":
                list2.append(name)
            break
        start_row += 1


def create_list_from_column_add_param(column_start, read_sheet, read_column, param_column, param):
    my_list = []
    acc = column_start
    while (read_sheet.cell(row=acc, column=read_column)).value:
        if (read_sheet.cell(row=acc, column=param_column)).value == param:
            my_list.append((read_sheet.cell(row=acc, column=read_column)).value)
        acc += 1
    return my_list


def list_of_dates(start_date, end_date):
    date_list = []
    date_1 = datetime.datetime.strptime(start_date, "%b-%d")
    date_2 = datetime.datetime.strptime(end_date, "%b-%d")
    delta_t = date_2 - date_1
    total_days = int((delta_t.total_seconds()) / 86400)
    for i in range(0, total_days+1):
        num_days = datetime.timedelta(days=i)
        date_to_print = date_1 + num_days
        date_list.append(date_mon_dd(date_to_print))
    return date_list



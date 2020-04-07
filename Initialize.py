import datetime
import openpyxl
import Functions as help

# Open Workbooks
Entry = openpyxl.load_workbook('BookForProgramDirector.xlsx')
Data = openpyxl.load_workbook('BookForPython.xlsx')

# Constants for Session Dates sheet
SESSION_DATES_SHEET = Entry["Session Dates"]
SESSION_COLUMN = 1
SESSION_START_COLUMN = 2
SESSION_END_COLUMN = 3
A1_ROW = 2
A2_ROW = 3
B1_ROW = 4
B2_ROW = 5
C_ROW = 6
A1_START = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_START_COLUMN)
A2_START = SESSION_DATES_SHEET.cell(row=A2_ROW, column=SESSION_START_COLUMN)
A1_END = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_END_COLUMN)


# Constants for Full Day, Half Day, and Duty sheets
FULL_DAY_SHEET = Entry["Full Day"]
HALF_DAY_SHEET = Entry["Half Day"]
DUTY_SCHEDULE_SHEET = Entry["Duty Schedule"]
STAFFING_SHEET = Entry["Official Staffing"]
DATE_COLUMN = 2
DATE_ROW = 2
NAME_COLUMN_A = 2
TYPE_COLUMN_A = 1

# Constants for Staff Roster
STAFF_ROSTER_SHEET = Entry["Staff Roster"]
NAME_COLUMN = 1
TYPE_COLUMN = 2
totalStaff = (STAFF_ROSTER_SHEET.cell(row=2, column=12)).value
typeList = ["JB", "JG", "BB", "BG", "IB", "IG", "SB", "SG", "SSB", "SSG", "LT", "SPEC", "TRIPPER"]


# Constants for Jinci Boys and Jinci Girls Rotation, Jinci Boys and Jinci Girls Schedule
JB_ROTATION_SHEET = Data["Jinci Boys Rotation"]
JG_ROTATION_SHEET = Data["Jinci Girls Rotation"]
JB_SCHEDULE_SHEET = Entry["Jinci Boys Schedule"]
JG_SCHEDULE_SHEET = Entry["Jinci Girls Schedule"]
ROTATION_COLUMN = 1
CABIN_COLUMN = 2
OFFSET_COLUMN = 3
CABIN1_ROW = 2
CABIN2_ROW = 3
CABIN3_ROW = 4
TotalCabinsJB = (JB_ROTATION_SHEET.cell(row=2, column=4)).value

# Create lists from rotation column
rotationJB = help.create_list_from_column(2, JB_ROTATION_SHEET, ROTATION_COLUMN)
rotationLengthJB = len(rotationJB)

rotationJG = help.create_list_from_column(2, JG_ROTATION_SHEET, ROTATION_COLUMN)
rotationLengthJG = len(rotationJG)

# Create Jinci Cabin Schedules from Rotation Sheet
help.create_jinci_schedules(2, JB_ROTATION_SHEET, JB_SCHEDULE_SHEET, rotationJB)
help.create_jinci_schedules(2, JG_ROTATION_SHEET, JG_SCHEDULE_SHEET, rotationJG)

# Take names from Staff Roster and print onto Duty Schedule Sheet
help.sort_names_by_type(STAFF_ROSTER_SHEET, 2, totalStaff, NAME_COLUMN, DUTY_SCHEDULE_SHEET, 3, NAME_COLUMN_A,
                        typeList, TYPE_COLUMN, TYPE_COLUMN_A)

# Take names from Staff Roster and print onto Official Staffing sheet
start_row = 2
while (STAFF_ROSTER_SHEET.cell(row=start_row, column=NAME_COLUMN)).value:
    name = (STAFF_ROSTER_SHEET.cell(row=start_row, column=NAME_COLUMN)).value
    STAFFING_SHEET.cell(row=start_row, column=NAME_COLUMN, value=name)
    start_row += 1

# Take dates from Session Dates and print onto Full Day Sheet, Half Day Sheet, and Duty Schedule Sheet
help.print_dates_to_column_from_table(A1_ROW, C_ROW, SESSION_START_COLUMN, SESSION_END_COLUMN, SESSION_DATES_SHEET,
                                      DATE_COLUMN, 2, FULL_DAY_SHEET)
help.print_dates_to_column_from_table(A1_ROW, C_ROW, SESSION_START_COLUMN, SESSION_END_COLUMN, SESSION_DATES_SHEET,
                                      DATE_COLUMN, 2, HALF_DAY_SHEET)
help.print_dates_to_row_from_table(A1_ROW, C_ROW, SESSION_START_COLUMN, SESSION_END_COLUMN, SESSION_DATES_SHEET,
                                   DATE_ROW, NAME_COLUMN_A + 1, DUTY_SCHEDULE_SHEET)


Entry.save("BookForProgramDirector.xlsx")

import openpyxl
import Functions

Entry = openpyxl.load_workbook('BookForProgramDirector.xlsx')
Data = openpyxl.load_workbook('BookForPython.xlsx')

DUTY_SCHEDULE_SHEET = Entry["Duty Schedule"]
FULL_DAY_SHEET = Entry["Full Day"]
HALF_DAY_SHEET = Entry["Half Day"]
SESSION_DATES_SHEET = Entry["Session Dates"]

B1_START = (SESSION_DATES_SHEET.cell(row=4, column=2)).value
C_START = (SESSION_DATES_SHEET.cell(row=6, column=2)).value

DUTY_SCHEDULE_NAME_COLUMN = 2
DAY_OFF_DATES_COLUMN = 2

END_DATE = B1_START

# Print A session day offs to full day and half day sheets

date_row_acc = 3
while (FULL_DAY_SHEET.cell(row=date_row_acc, column=DAY_OFF_DATES_COLUMN)).value != END_DATE:

    date_column_acc = date_row_acc + 1

    name_row_acc = 3
    half_day_off_slot_acc = 3
    full_day_off_slot_acc = 3
    while (DUTY_SCHEDULE_SHEET.cell(row=name_row_acc, column=DUTY_SCHEDULE_NAME_COLUMN)).value:

        name = (DUTY_SCHEDULE_SHEET.cell(row=name_row_acc, column=DUTY_SCHEDULE_NAME_COLUMN)).value

        if (DUTY_SCHEDULE_SHEET.cell(row=name_row_acc, column=date_column_acc)).value == "F":
            FULL_DAY_SHEET.cell(row=date_row_acc, column=full_day_off_slot_acc, value=name)
            full_day_off_slot_acc += 1

        elif (DUTY_SCHEDULE_SHEET.cell(row=name_row_acc, column=date_column_acc)).value in ["H", "T/H"]:
            HALF_DAY_SHEET.cell(row=date_row_acc, column=half_day_off_slot_acc, value=name)
            half_day_off_slot_acc += 1

        name_row_acc += 1

    date_row_acc += 1

Entry.save('BookForProgramDirector.xlsx')

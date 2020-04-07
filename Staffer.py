# THIS IS AN OLD VERSION OF THE PROGRAM. DOES NOT WORK WITH THE CURRENT SETUP
# Loads Excel Workbook and defines sheet names
import random
import datetime
import tkinter as tk
from tkinter import *
from openpyxl import load_workbook


def name_jeff():
    print("Name Jeff")


def set_current_date():
    date_and_time = datetime.datetime.now()
    day_of_month = date_and_time.strftime("%d")
    month = date_and_time.strftime("%b")
    global currentDate
    currentDate = month + "-" + day_of_month
    print(currentDate)


def use_input_date():
    global currentDate
    currentDate = dateEntryBox.get()
    print("The date is %s" % currentDate)


badActivityList = []


def add_bad_activity():
    bad_activity = enterBadActivity.get()
    badActivityList.append(bad_activity)
    print(badActivityList)
    enterBadActivity.delete(0, tk.END)
    activityString.set(badActivityList)


root = tk.Tk()
frame = tk.Frame(root)
frame.pack()
dateButton = tk.Button(frame, text="Or you can use today's date", fg="red", command=lambda: [set_current_date(), root.destroy()])
dateButton.grid(row=2, column=0)
quitProgram = tk.Button(frame, text="QUIT", fg="red", command=quit)
quitProgram.grid(row=2, column=3)
fillerLabel = tk.Label(frame, text="-")
fillerLabel.grid(row=1)
showDate = tk.Button(frame, text='Submit', command=lambda: [use_input_date(), root.destroy()])
showDate.grid(row=0, column=2)
dateEntryLabel = tk.Label(frame, text="Enter a date (Mon-xx):")
dateEntryLabel.grid(row=0, column=0)
dateEntryBox = tk.Entry(frame)
dateEntryBox.grid(row=0, column=1)
root.mainloop()

root2 = tk.Tk()
frame2 = tk.Frame(root2)
frame2.pack()
badActivities = tk.Label(frame2, text="Enter Activities that you don't want to run:")
badActivities.grid(row=0, column=0)
activityString = StringVar(badActivityList)
enterBadActivity = tk.Entry(frame2)
enterBadActivity.grid(row=0, column=1)
notRunning = tk.Label(frame2, text="Not running the following activities:")
notRunning.grid(row=2,column=0)
fillerLabel = tk.Label(frame2, text="-")
fillerLabel.grid(row=1)
listBadActivities = tk.Label(frame2, textvariable=activityString)
listBadActivities.grid(row=2, column=1)
submit = tk.Button(frame2, text='Submit Activity', command=lambda: [add_bad_activity(), root2.update_idletasks()])
submit.grid(row=0, column=2)
runProgram = tk.Button(frame2, text='Run Program!', command=root2.destroy)
runProgram.grid(row=3, column=0)
root2.mainloop()

Entry = load_workbook('BookForProgramDirector')
Data = load_workbook('BookForPython')

spares = True

# For both morning and afternoon
for time in range(0, 2):
    if time == 0:
        morning = True
    else:
        morning = False
    if morning:
        staffSheet = Entry["Morning Staff"]
        activitySheet = Entry["Morning Activities"]
    else:
        staffSheet = Entry["Afternoon Staff"]
        activitySheet = Entry["Afternoon Activities"]
    FULL_DAY_SHEET = Entry["Full Day"]
    HALF_DAY_SHEET = Entry["Half Day"]
    STAFFING_SHEET = Entry["Official Staffing"]
    tripSheet = Entry["Trip"]

    # Define Constants

    numberOfActivities = 25
    numberOfStaffMembers = 81
    ropesRow = 7
    dayOffColumn = 2
    specColumn = 3
    NLColumn = 4
    crossColumn = 5
    ropesTrainedColumn = 6
    workingAtColumn = 7
    firstPreferenceColumn = 8
    secondPreferenceColumn = 9
    thirdPreferenceColumn = 10
    tripColumn = 11
    activityNameColumn = 1
    minStaffColumn = 2
    numStaffColumn = 3
    maxStaffColumn = 4
    ropesColumn = 5
    waterfrontColumn = 6
    aptlyStaffed = 7

    # Clear staffing
    if morning:
        for staff in range(2, numberOfStaffMembers + 2):
            for period in range(2, 7):
                STAFFING_SHEET.cell(row=staff, column=period, value="---")

    # Creates a list with a value for each staff member
    staffList = []
    for number in range(2, numberOfStaffMembers + 2):
        staffList.append(number)

    random.shuffle(staffList)
    print(staffList)

    # Set days off
    for day in range(2, 60):
        fullDay = FULL_DAY_SHEET.cell(row=day, column=1)
        halfDay = HALF_DAY_SHEET.cell(row=day, column=1)
        if fullDay.value in [currentDate]:
            for staff in range(2, numberOfStaffMembers + 2):
                staffName = staffSheet.cell(row=staff, column=1)
                for dayOffParticipant in range(2, 21):
                    if morning:
                        onFullDay = FULL_DAY_SHEET.cell(row=day - 1, column=dayOffParticipant)
                        onHalfDay = HALF_DAY_SHEET.cell(row=day - 1, column=dayOffParticipant)
                    else:
                        onFullDay = FULL_DAY_SHEET.cell(row=day, column=dayOffParticipant)
                        onHalfDay = HALF_DAY_SHEET.cell(row=1, column=1)
                    if (onFullDay.value in [staffName.value]) or (onHalfDay.value in [staffName.value]):
                        staffSheet.cell(row=staff, column=dayOffColumn, value="Yes")
                        break
                    else:
                        staffSheet.cell(row=staff, column=dayOffColumn, value="No")
            break

    # Set current number of staff at every activity to 0
    for y in range(2, numberOfActivities + 2):
        activitySheet.cell(row=y, column=numStaffColumn, value=0)

    # Set "working at" status of all staff to "Not Staffed"
    for y in range(2, numberOfStaffMembers + 2):
        staffSheet.cell(row=y, column=workingAtColumn, value="Not Staffed")
    for y in range(2, numberOfActivities + 2):
        activitySheet.cell(row=y, column=aptlyStaffed, value="GOOD")

    # Who is on trip
    for day in range(2, 60):
        tripDay = tripSheet.cell(row=1, column=day)
        if tripDay.value in [currentDate]:
            for staff in range(2, numberOfStaffMembers + 2):
                onTrip = tripSheet.cell(row=staff, column=day)
                if onTrip.value in ["Yes"]:
                    staffSheet.cell(row=staff, column=tripColumn, value="Yes")
            break

    def define_activity_attributes(activity):
        global numStaff
        global minStaff
        global maxStaff
        global activityName
        global waterFront
        global ropesActivity
        numStaff = activitySheet.cell(row=activity, column=numStaffColumn)
        minStaff = activitySheet.cell(row=activity, column=minStaffColumn)
        maxStaff = activitySheet.cell(row=activity, column=maxStaffColumn)
        activityName = activitySheet.cell(row=activity, column=activityNameColumn)
        waterFront = activitySheet.cell(row=activity, column=waterfrontColumn)
        ropesActivity = activitySheet.cell(row=activity, column=ropesColumn)

    def loop_through_activities(preference, mode):
        for y in range(2, numberOfActivities + 2):
            activity_name = activitySheet.cell(row=y, column=activityNameColumn)
            if activity_name.value not in badActivityList:
                num_staff = activitySheet.cell(row=y, column=numStaffColumn)
                min_staff = activitySheet.cell(row=y, column=minStaffColumn)
                max_staff = activitySheet.cell(row=y, column=maxStaffColumn)
                true_num_staff = num_staff.value
                water_front = activitySheet.cell(row=y, column=waterfrontColumn)
                ropes_activity = activitySheet.cell(row=y, column=ropesColumn)
                swim_no_prob = bool(
                    (water_front.value in ["No"]) or ((water_front.value in ["Yes"]) and (cross.value in ["Yes"])))
                ropes_no_prob = bool((ropes_activity.value not in ["Yes"]) or (
                        (ropes_activity.value in ["Yes"]) and (ropesTrained.value in ["Yes"])))
                matches_preference = bool(activity_name.value == preference.value)
                below_max = bool(num_staff.value < max_staff.value)
                below_min = bool(num_staff.value < min_staff.value)
                mode_0 = bool(swim_no_prob and ropes_no_prob and below_min and (mode == 0))
                mode_1 = bool(swim_no_prob and ropes_no_prob and matches_preference and below_max and (mode == 1))
                mode_2 = bool(swim_no_prob and ropes_no_prob and below_max and (mode == 2))
                if mode_0 or mode_1 or mode_2:
                    staffSheet.cell(row=p, column=workingAtColumn, value=activity_name.value)
                    true_num_staff = true_num_staff + 1
                    activitySheet.cell(row=y, column=numStaffColumn, value=true_num_staff)
                    break
        return


    def define_staff_attributes():
        global dayOff
        global cross
        global spec
        global working
        global preference1
        global preference2
        global preference3
        global ropesTrained
        global NL
        global trip
        dayOff = staffSheet.cell(row=p, column=dayOffColumn)
        cross = staffSheet.cell(row=p, column=crossColumn)
        spec = staffSheet.cell(row=p, column=specColumn)
        working = staffSheet.cell(row=p, column=workingAtColumn)
        preference1 = staffSheet.cell(row=p, column=firstPreferenceColumn)
        preference2 = staffSheet.cell(row=p, column=secondPreferenceColumn)
        preference3 = staffSheet.cell(row=p, column=thirdPreferenceColumn)
        ropesTrained = staffSheet.cell(row=p, column=ropesTrainedColumn)
        NL = staffSheet.cell(row=p, column=NLColumn)
        trip = staffSheet.cell(row=p, column=tripColumn)


    # Staff all specs at their specialized activity
    for x in range(0, numberOfStaffMembers):
        p = staffList[x]
        define_staff_attributes()
        if spec.value in ["Yes"] or (NL.value in ["Yes"]):
            loop_through_activities(preference=preference1, mode=1)
        if dayOff.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Day Off")
        if trip.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Trip")

    # Staff everyone else at their first choice if possible. Then go onto second, third, etc.
    for x in range(0, numberOfStaffMembers):
        p = staffList[x]

        define_staff_attributes()
        if dayOff.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Day Off")
        if trip.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Trip")
        elif spec.value not in ["Yes"] and (NL.value not in ["Yes"]):
            loop_through_activities(preference=preference1, mode=1)
            if working.value in ["Not Staffed"]:
                loop_through_activities(preference=preference2, mode=1)
            if working.value in ["Not Staffed"]:
                loop_through_activities(preference=preference3, mode=1)

    # Place remaining staff at activities below minimum staff; make sure everyone is staffed
    for x in range(0, numberOfStaffMembers):
        p = staffList[x]
        define_staff_attributes()
        if trip.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Trip")
        if dayOff.value in ["Yes"]:
            staffSheet.cell(row=p, column=workingAtColumn, value="Day Off")
        if working.value in ["Not Staffed"]:
            loop_through_activities(preference=working, mode=0)
        if working.value in ["Not Staffed"]:
            loop_through_activities(preference=working, mode=2)

    # Make sure ropes is adequately staffed
    define_activity_attributes(activity=ropesRow)
    if activityName.value not in badActivityList:
        trueNumStaff = numStaff.value
        for x in range(0, numberOfStaffMembers):
            if numStaff.value >= minStaff.value:
                break
            p = staffList[numberOfStaffMembers - x - 1]
            define_staff_attributes()
            if ropesActivity.value in ["Yes"]:
                for z in range(2, numberOfActivities + 2):
                    numStaff2 = activitySheet.cell(row=z, column=numStaffColumn)
                    minStaff2 = activitySheet.cell(row=z, column=minStaffColumn)
                    activityName2 = activitySheet.cell(row=z, column=activityNameColumn)
                    waterFront2 = activitySheet.cell(row=z, column=waterfrontColumn)
                    trueNumStaff2 = numStaff2.value
                    weGood = bool((ropesTrained.value in ["Yes"]) and (numStaff2.value > 0) and
                                  (working.value not in ["Challenge"]))
                    if (working.value == activityName2.value) and weGood:
                        staffSheet.cell(row=p, column=workingAtColumn, value=activityName.value)
                        trueNumStaff = trueNumStaff + 1
                        trueNumStaff2 = trueNumStaff2 - 1
                        activitySheet.cell(row=ropesRow, column=numStaffColumn, value=trueNumStaff)
                        activitySheet.cell(row=z, column=numStaffColumn, value=trueNumStaff2)
                        break

    # For any other activities still below minimum staff, grab staff from activities that are above minimum staff
    # if the understaffed activity matches one of their preferences
    for y in range(2, numberOfActivities + 2):
        define_activity_attributes(activity=y)
        if activityName.value not in badActivityList:
            trueNumStaff = numStaff.value

            for x in range(0, numberOfStaffMembers):
                if numStaff.value >= minStaff.value:
                    break
                p = staffList[numberOfStaffMembers - x - 1]
                define_staff_attributes()
                matchesPreference = bool(
                    activityName.value == (preference1.value or preference2.value or preference3.value))
                if spec.value not in ["Yes"] and (NL.value not in ["Yes"]) and matchesPreference:
                    for z in range(2, numberOfActivities + 2):
                        activityName2 = activitySheet.cell(row=z, column=activityNameColumn)
                        if activityName2.value not in badActivityList:
                            numStaff2 = activitySheet.cell(row=z, column=numStaffColumn)
                            minStaff2 = activitySheet.cell(row=z, column=minStaffColumn)
                            waterFront2 = activitySheet.cell(row=z, column=waterfrontColumn)
                            trueNumStaff2 = numStaff2.value
                            if (working.value == activityName2.value) and (
                                    numStaff2.value > minStaff2.value) and (
                                    (waterFront.value in ["No"]) or (
                                    (waterFront.value in ["Yes"]) and (cross.value in ["Yes"]))) and (
                                    (ropesActivity.value not in ["Yes"]) or
                                    ((ropesActivity.value in ["Yes"]) and
                                     (ropesTrained.value in ["Yes"])
                                     )):
                                staffSheet.cell(row=p, column=workingAtColumn, value=activityName.value)
                                trueNumStaff = trueNumStaff + 1
                                trueNumStaff2 = trueNumStaff2 - 1
                                activitySheet.cell(row=y, column=numStaffColumn, value=trueNumStaff)
                                activitySheet.cell(row=z, column=numStaffColumn, value=trueNumStaff2)
                                break

    # For any other activities still below minimum staff, grab staff from activities that are above minimum staff
    for y in range(2, numberOfActivities + 2):
        define_activity_attributes(activity=y)
        trueNumStaff = numStaff.value
        if activityName.value not in badActivityList:
            for x in range(0, numberOfStaffMembers):
                if numStaff.value >= minStaff.value:
                    break
                p = staffList[numberOfStaffMembers - x - 1]
                define_staff_attributes()
                if spec.value not in ["Yes"] and (NL.value not in ["Yes"]):
                    for z in range(2, numberOfActivities + 2):
                        activityName2 = activitySheet.cell(row=z, column=activityNameColumn)
                        if activityName2.value not in badActivityList:
                            numStaff2 = activitySheet.cell(row=z, column=numStaffColumn)
                            minStaff2 = activitySheet.cell(row=z, column=minStaffColumn)
                            waterFront2 = activitySheet.cell(row=z, column=waterfrontColumn)
                            trueNumStaff2 = numStaff2.value
                            if (working.value == activityName2.value) and (
                                    numStaff2.value > minStaff2.value) and (
                                    (waterFront.value in ["No"]) or (
                                    (waterFront.value in ["Yes"]) and (cross.value in ["Yes"]))) and (
                                    (ropesActivity.value not in ["Yes"]) or
                                    ((ropesActivity.value in ["Yes"]) and
                                     (ropesTrained.value in ["Yes"])
                                    )):
                                staffSheet.cell(row=p, column=workingAtColumn, value=activityName.value)
                                trueNumStaff = trueNumStaff + 1
                                trueNumStaff2 = trueNumStaff2 - 1
                                activitySheet.cell(row=y, column=numStaffColumn, value=trueNumStaff)
                                activitySheet.cell(row=z, column=numStaffColumn, value=trueNumStaff2)
                                break

    # Let user know if an activity is understaffed
    for act in range(2, numberOfActivities + 2):
        if activityName.value not in badActivityList:
            numStaff = activitySheet.cell(row=act, column=numStaffColumn)
            minStaff = activitySheet.cell(row=act, column=minStaffColumn)
            if numStaff.value < minStaff.value:
                activitySheet.cell(row=act, column=aptlyStaffed, value="NEEDS MORE STAFF!!!")

    # Print the amount of staff that got (or didn't get) their preference
    firstPreferenceTally = 0
    secondPreferenceTally = 0
    thirdPreferenceTally = 0
    dayOffTally = 0

    for p in range(2, numberOfStaffMembers + 2):
        define_staff_attributes()
        if working.value in ["Day Off"]:
            dayOffTally += 1
        elif working.value == preference1.value:
            firstPreferenceTally += 1
        elif working.value == preference2.value:
            secondPreferenceTally += 1
        elif working.value == preference3.value:
            thirdPreferenceTally += 1

    notStoked = numberOfStaffMembers - (
                firstPreferenceTally + secondPreferenceTally + thirdPreferenceTally + dayOffTally)
    print("First preference:", firstPreferenceTally, "\nSecond Preference:", secondPreferenceTally,
          "\nThird Preference", thirdPreferenceTally, "\nNot stoked:", notStoked)
    staffSheet.cell(row=2, column=14, value=firstPreferenceTally)
    staffSheet.cell(row=3, column=14, value=secondPreferenceTally)
    staffSheet.cell(row=4, column=14, value=thirdPreferenceTally)
    staffSheet.cell(row=5, column=14, value=notStoked)

    for staffMember in range(2, numberOfStaffMembers + 2):
        workingAt = staffSheet.cell(row=staffMember, column=workingAtColumn)
        if morning:
            for period in range(2, 5):
                STAFFING_SHEET.cell(row=staffMember, column=period, value=workingAt.value)
        else:
            for period in range(5, 7):
                STAFFING_SHEET.cell(row=staffMember, column=period, value=workingAt.value)

    # Give spares
    for act in range(2, numberOfActivities + 2):
        define_activity_attributes(activity=act)
        if activityName.value not in badActivityList:
            if numStaff.value >= minStaff.value + 2:
                gotSpare = STAFFING_SHEET.cell(row=1, column=1)
                for period in range(1, 3):
                    if morning:
                        p = period * 2
                    else:
                        p = period + 4
                    for x in range(0, numberOfStaffMembers):
                        staff = staffList[x]
                        staffName = STAFFING_SHEET.cell(row=staff, column=1)
                        working = STAFFING_SHEET.cell(row=staff, column=p)
                        if (working.value == activityName.value) and (staffName.value != gotSpare.value):
                            gotSpare = STAFFING_SHEET.cell(row=staff, column=1)
                            STAFFING_SHEET.cell(row=staff, column=p, value="Spare")
                            break

    Entry.save('NewStaffList.xlsx')

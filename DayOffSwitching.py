import datetime
import openpyxl
import Functions
import random
import array

# Load workbooks
Entry = openpyxl.load_workbook('BookForProgramDirector.xlsx')
Data = openpyxl.load_workbook('BookForPython.xlsx')

# Constants for Session Dates sheet
SESSION_DATES_SHEET = Entry["Session Dates"]
SESSION_COLUMN = 1
SESSION_START_COLUMN = 2
SESSION_END_COLUMN = 3
A1_ROW = 2
A2_ROW = 3
B1_ROW = 4
B2_ROW = 5
C_ROW = 6
A1_START = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_START_COLUMN)
A2_START = SESSION_DATES_SHEET.cell(row=A2_ROW, column=SESSION_START_COLUMN)
A1_END = SESSION_DATES_SHEET.cell(row=A1_ROW, column=SESSION_END_COLUMN)
A2_END = SESSION_DATES_SHEET.cell(row=A2_ROW, column=SESSION_END_COLUMN)
B1_START = SESSION_DATES_SHEET.cell(row=B1_ROW, column=SESSION_START_COLUMN)
A1_LIST = Functions.list_of_dates(A1_START.value, A1_END.value)
A2_LIST = Functions.list_of_dates(A2_START.value, A2_END.value)
SESSIONS = [A1_LIST, A2_LIST]

# Constants for Duty Schedule
DUTY_SCHEDULE_SHEET = Entry["Duty Schedule"]
DATE_ROW = 2
DUTY_NAME_COLUMN = 2
FIRST_NAME_ROW = 3  # The first row containing a staff's name

DAY_OFF_SWITCHES_SHEET = Entry["Day Off Switches"]

# The following three lists are of the form [[Staff, Date Switched Off, Session, "H/F"], repeat...]
SWITCHED_FOR_STAFF = []
SWITCHED_FOR_NLS = []
SWITCHED_FOR_ROPES = []
NUM_BLACK_OUT_DAYS = 3
FINAL_SWITCHES_LIST = []

# This list will store the final amounts of staff for each date
staff_numbers_by_date = []

# Constants for Staff Roster Sheet
STAFF_ROSTER_SHEET = Entry["Staff Roster"]
ROSTER_NAME_COLUMN = 1
TYPE_COLUMN = 2
STAFF_NLS_COLUMN = 3
CROSS_COLUMN = 4
ROPES_TRAINED_COLUMN = 5
STAFF_LIST = Functions.create_list_from_column(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN)
# Create lists for staff of every kind:
NLS_STAFF_LIST = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN,STAFF_NLS_COLUMN,
                                                             "Y")
ROPES_STAFF_LIST = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN,
                                                               ROPES_TRAINED_COLUMN, "Y")
LEADERSHIP_TEAM = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN,
                                                              "LT")
SPECS = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "SPEC")
TRIPPERS = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN,
                                                       "TRIPPER")
JB = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "JB")
JG = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "JG")
BB = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "BB")
BG = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "BG")
IB = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "IB")
IG = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "IG")
SB = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "SB")
SG = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "SG")
SSB = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "SSB")
SSG = Functions.create_list_from_column_add_param(2, STAFF_ROSTER_SHEET, ROSTER_NAME_COLUMN, TYPE_COLUMN, "SSG")
ALL_SECTIONS = [JB, JG, BB, BG, IB, IG, SB, SG, SSB, SSG]


print(LEADERSHIP_TEAM, SPECS, TRIPPERS)
# Constants for Activity Numbers Sheet
ACTIVITY_NUMBERS_SHEET = Data["Activity Numbers"]
ACTIVITY_COLUMN = 1
MIN_STAFF_COLUMN = 2
NUM_STAFF_COLUMN = 3
MAX_STAFF_COLUMN = 4
ROPES_COLUMN = 5
WATERFRONT_COLUMN = 6
ACTIVITY_NLS_COLUMN = 7
FIRST_DATE_TO_CHECK = "Jul"

# Calculates the minimum number of staff, ropes staff, and NLS staff required; adds a buffer of 4
BUFFER = 4
totalNLSRequired = Functions.sum_column_while_value(ACTIVITY_NUMBERS_SHEET, 2, ACTIVITY_NLS_COLUMN, ACTIVITY_COLUMN) + \
                   BUFFER
totalStaffRequired = Functions.sum_column_while_value(ACTIVITY_NUMBERS_SHEET, 2, MIN_STAFF_COLUMN, ACTIVITY_COLUMN) + \
                     BUFFER
totalRopesRequired = 12 + BUFFER


# Set start date and end date. This will depend on the session. Also create a list of times of day: [Morning, Afternoon]
start_date = A1_START.value
end_date = B1_START.value
TIME_OF_DAY = ["Morning", "Afternoon"]

# Get start column for given start date
acc = 1
while (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).value:
    if (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).value == start_date:
        start_column = (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).column
        break
    acc += 1

# Go through all dates for session and check if there is adequate amount of staff
while (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=start_column)).value != end_date:
    date = (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=start_column)).value

    # If the date is in A1, we're in A1. If the date is in A2, we're in A2. Etc...
    for session in SESSIONS:
        if date in session:
            current_session = session

    # Create empty lists for the given date denoting staff that are present. Index 0 is afternoon, Index 1 is next morn
    present = [[], []]
    presentWithNLS = [[], []]
    presentRopesTrained = [[], []]

    # Go down all staff in the duty schedule
    acc = FIRST_NAME_ROW
    while (DUTY_SCHEDULE_SHEET.cell(row=acc, column=DUTY_NAME_COLUMN)).value:

        # If a staff is NOT on day off in afternoon, add them to the afternoon present list.
        # If they also happen to have their NLS or Ropes training, add them to those respective lists as well

        if (DUTY_SCHEDULE_SHEET.cell(row=acc, column=start_column)).value not in ["T", "F", "T/H"]:
            name = (DUTY_SCHEDULE_SHEET.cell(row=acc, column=DUTY_NAME_COLUMN)).value
            present[0].append(name)
            Functions.add_names_to_lists_if(presentWithNLS[0], presentRopesTrained[0], STAFF_NLS_COLUMN,
                                            ROPES_TRAINED_COLUMN,
                                            name, ROSTER_NAME_COLUMN, STAFF_ROSTER_SHEET)

        # If a staff is NOT on day off the next morning, add them to the morning present list.
        # If they also happen to have their NLS or Ropes training, add them to those respective lists as well
        if ((DUTY_SCHEDULE_SHEET.cell(row=acc, column=start_column + 1)).value not in ["T"] and
                (DUTY_SCHEDULE_SHEET.cell(row=acc, column=start_column)).value not in ["H", "F", "T/H"]):
            name = (DUTY_SCHEDULE_SHEET.cell(row=acc, column=DUTY_NAME_COLUMN)).value
            present[1].append(name)
            Functions.add_names_to_lists_if(presentWithNLS[1], presentRopesTrained[1], STAFF_NLS_COLUMN,
                                            ROPES_TRAINED_COLUMN,
                                            name, ROSTER_NAME_COLUMN, STAFF_ROSTER_SHEET)

        acc += 1

    # Once we've gone through all staff, take the length of each of these lists created for the given date
    # Index 0 is afternoon, Index 1 is next morning
    totalStaffAvailable = [len(present[0]), len(present[1])]
    totalNLSAvailable = [len(presentWithNLS[0]), len(presentWithNLS[1])]
    totalRopesAvailable = [len(presentRopesTrained[0]), len(presentRopesTrained[1])]

    def day_off_switcher(attribute, available, required, total_available, total_required, staff_list, switched):

        # This is what we do when there aren't enough staff with a certain attribute in the afternoon or next morning
        if available[0] < required or available[1] < required:
            if available[0] < required:
                print("Not enough", attribute, "staff in afternoon")
            if available[1] < required:
                print("Not enough", attribute, "staff next morning")

            # The following process removes staff from day off for the given day and adds them to list of staff
            # members who need to be added to another day off

            # Put staff list in random order
            random.shuffle(staff_list)

            # For all staff in staff list
            for staff in range(0, len(staff_list)):

                # Search through every row of duty schedule
                start_row = FIRST_NAME_ROW
                while (DUTY_SCHEDULE_SHEET.cell(row=start_row, column=DUTY_NAME_COLUMN)).value:

                    # If staff in this row of duty schedule is the staff at the current spot in staff list
                    # And they're not on the leadership team / Spec / Tripper
                    if (DUTY_SCHEDULE_SHEET.cell(row=start_row, column=DUTY_NAME_COLUMN)).value == staff_list[staff] \
                            and (staff_list[staff] not in (LEADERSHIP_TEAM + SPECS + TRIPPERS)):

                        # If we need more afternoon staff and they're on full day
                        if (available[0] < required) and \
                                ((DUTY_SCHEDULE_SHEET.cell(row=start_row, column=start_column)).value in ["F"]):

                            # Switch this staff member and update daily availability
                            staff_name = (DUTY_SCHEDULE_SHEET.cell(row=start_row, column=DUTY_NAME_COLUMN)).value
                            print(staff_name, " switched off", date)
                            switched.append([staff_name, date, current_session, "F"])
                            available[0] += 1
                            total_available[0] += 1
                            available[1] += 1
                            total_available[1] += 1

                            break

                        # Otherwise if we need more morning staff and they're on half day
                        elif (available[1] < required) and \
                                ((DUTY_SCHEDULE_SHEET.cell(row=start_row, column=start_column)).value in ["H"]):

                            # Switch this staff member and update daily availability
                            staff_name = (DUTY_SCHEDULE_SHEET.cell(row=start_row, column=DUTY_NAME_COLUMN)).value
                            print(staff_name, " switched off", date)
                            switched.append([staff_name, date, current_session, "H"])
                            available[1] += 1
                            total_available[1] += 1

                        # Otherwise if we need more morning staff and they're on full day
                        elif (available[1] < required) and \
                                ((DUTY_SCHEDULE_SHEET.cell(row=start_row, column=start_column)).value in ["F"]):

                            # Switch this staff member and update daily availability
                            staff_name = (DUTY_SCHEDULE_SHEET.cell(row=start_row, column=DUTY_NAME_COLUMN)).value
                            print(staff_name, " switched off", date)
                            switched.append([staff_name, date, current_session, "F"])
                            available[0] += 1
                            total_available[0] += 1
                            available[1] += 1
                            total_available[1] += 1

                    start_row += 1

                if available[0] >= required and available[1] >= required:
                    break

        # Otherwise if there are more than enough staff with this attribute, switch staff ONTO this day off
        else:
            print("Sufficient", attribute, "staff")

            # If it's not a blackout day
            if (current_session.index(date) < len(current_session) - 3) and (current_session.index(date) > 2):

                # While there are more than enough staff in the afternoon and next morning with the given attribute:
                while available[0] > required and total_available[0] > total_required and \
                        available[1] > required and total_available[1] > total_required:

                    # If there are no more staff to be switched, quit the loop
                    if len(switched) == 0:
                        break

                    initial_length = len(switched)

                    # For every staff that needs to be switched onto a day off
                    for staff in range(0, len(switched)):

                        # Don't switch a staff onto the same day they were switched off of
                        # Make sure switches remain within one session. Can't switch an A1 day off to A2
                        # If they're on a full day

                        if switched[staff][1] != date and switched[staff][2] == current_session and \
                                switched[staff][3] == "F":

                            # Get this staff member's row in the duty schedule
                            row = FIRST_NAME_ROW
                            while (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value:
                                if (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value == \
                                        switched[staff][0]:
                                    break
                                row += 1

                            # Get this staff member's section
                            for section in ALL_SECTIONS:
                                if switched[staff][0] in section:
                                    break

                            # Check how many staff in this section are on day off or trip
                            absent = 0
                            for staff_member in section:
                                row = FIRST_NAME_ROW
                                while (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value:
                                    if (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value == \
                                            staff_member:
                                        break
                                    row += 1

                                # If they're on day off or trip, add them to absent list
                                if (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column)).value in \
                                        ["H", "F", "T"]:
                                    absent += 1

                            # Don't allow two days off in a row
                            # Don't switch to a day when they're already off or when they're on trip
                            # Don't allow more than half of staff from a section to be gone
                            if (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column - 1)).value not in ["H", "F"] \
                                    and (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column)).value \
                                    not in ["H", "F", "T"] and \
                                    (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column + 1)).value \
                                    not in ["H", "F", "T"] and \
                                    absent < (len(section) / 2):
                                print(switched[staff][0], "switched from", switched[staff][1], "to", date)
                                temp_string = (switched[staff][0] + " switched from " + switched[staff][1] + " to " + date)
                                FINAL_SWITCHES_LIST.append(temp_string)
                                del switched[staff]
                                available[0] -= 1
                                total_available[0] -= 1
                                available[1] -= 1
                                total_available[1] -= 1
                                break

                    if len(switched) == initial_length:
                        break

                # If, after this, there are still excess staff in the morning, switch people onto half day
                while available[1] > required and total_available[0] > total_required:

                    if len(switched) == 0:
                        break

                    initial_length = len(switched)

                    # For every staff that needs to be switched onto a day off
                    for staff in range(0, len(switched)):

                        # Don't switch a staff onto the same day they were switched off of
                        # Make sure switches remain within one session. Can't switch an A1 day off to A2
                        # If they're on a full day

                        if switched[staff][1] != date and switched[staff][2] == current_session and \
                                switched[staff][3] == "H":

                            # Get this staff member's row in the duty schedule
                            row = FIRST_NAME_ROW
                            while (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value:
                                if (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value == \
                                        switched[staff][0]:
                                    break
                                row += 1

                            # Get this staff member's section
                            for section in ALL_SECTIONS:
                                if switched[staff][0] in section:
                                    break

                            # Check how many staff in this section are on day off or trip
                            absent = 0
                            for staff_member in section:
                                row = FIRST_NAME_ROW
                                while (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value:
                                    if (DUTY_SCHEDULE_SHEET.cell(row=row, column=DUTY_NAME_COLUMN)).value == \
                                            staff_member:
                                        break
                                    row += 1

                                # If they're on day off or trip, add them to absent list
                                if (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column)).value in \
                                        ["H", "F", "T"]:
                                    absent += 1

                            # Don't allow two days off in a row
                            # Don't switch to a day when they're already off or when they're on trip
                            # Don't allow more than half of staff from a section to be gone
                            if (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column - 1)).value not in ["H", "F"] \
                                    and (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column)).value \
                                    not in ["H", "F", "T"] and \
                                    (DUTY_SCHEDULE_SHEET.cell(row=row, column=start_column + 1)).value \
                                    not in ["H", "F", "T"] and \
                                    absent < (len(section) / 2):
                                print(switched[staff][0], "switched from", switched[staff][1], "to", date)
                                temp_string = (switched[staff][0] + " switched from " + switched[staff][1] + " to " + date)
                                FINAL_SWITCHES_LIST.append(temp_string)
                                del switched[staff]
                                available[1] -= 1
                                total_available[1] -= 1
                                break

                    if len(switched) == initial_length:
                        break

        return available, total_available, staff_list, switched


    print(date)

    # Carry out switches due to NLS
    totalNLSAvailable, totalStaffAvailable, NLS_STAFF_LIST, SWITCHED_FOR_NLS = \
        day_off_switcher("NLS", totalNLSAvailable, totalNLSRequired, totalStaffAvailable, totalStaffRequired,
                         NLS_STAFF_LIST, SWITCHED_FOR_NLS)

    # Carry out switches due to Ropes
    totalRopesAvailable, totalStaffAvailable, ROPES_STAFF_LIST, SWITCHED_FOR_ROPES = \
        day_off_switcher("Ropes", totalRopesAvailable, totalRopesRequired, totalStaffAvailable, totalStaffRequired,
                         ROPES_STAFF_LIST, SWITCHED_FOR_ROPES)

    # Carry out switches due to total staff
    totalStaffAvailable, dummy, STAFF_LIST, SWITCHED_FOR_STAFF = \
        day_off_switcher("", totalStaffAvailable, totalStaffRequired, totalStaffAvailable, totalStaffRequired,
                         STAFF_LIST, SWITCHED_FOR_STAFF)

    # Store the final staff numbers for this date
    staff_numbers_by_date.append([date, totalStaffAvailable, totalNLSAvailable, totalRopesAvailable])

    start_column += 1

print(staff_numbers_by_date)
print(totalStaffRequired)
# Once we've gone through all of the dates in the session, there may still be some staff that haven't been placed on a
# new day off. If so, we start from the beginning of the dates again:

print(SWITCHED_FOR_STAFF)

# Reset start column:
acc = 1
while (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).value:
    if (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).value == start_date:
        start_column = (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=acc)).column
        break
    acc += 1


# Go through dates again and perform remaining switches
index = 0
while len(SWITCHED_FOR_STAFF + SWITCHED_FOR_ROPES + SWITCHED_FOR_NLS) > 0 and \
        (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=start_column)).value != end_date:

    date = (DUTY_SCHEDULE_SHEET.cell(row=DATE_ROW, column=start_column)).value

    # If the date is in A1, we're in A1. If the date is in A2, we're in A2. Etc...
    for session in SESSIONS:
        if date in session:
            current_session = session
    print(date)
    # Carry out switches due to NLS
    totalNLSAvailable, totalStaffAvailable, NLS_STAFF_LIST, SWITCHED_FOR_NLS = \
        day_off_switcher("NLS", staff_numbers_by_date[index][2], totalNLSRequired, staff_numbers_by_date[index][1],
                         totalStaffRequired, NLS_STAFF_LIST, SWITCHED_FOR_NLS)

    # Carry out switches due to Ropes
    totalRopesAvailable, totalStaffAvailable, ROPES_STAFF_LIST, SWITCHED_FOR_ROPES = \
        day_off_switcher("Ropes", staff_numbers_by_date[index][3], totalRopesRequired, staff_numbers_by_date[index][1],
                         totalStaffRequired, ROPES_STAFF_LIST, SWITCHED_FOR_ROPES)

    # Carry out switches due to total staff
    totalStaffAvailable, dummy, STAFF_LIST, SWITCHED_FOR_STAFF = \
        day_off_switcher("", staff_numbers_by_date[index][1], totalStaffRequired, staff_numbers_by_date[index][1],
                         totalStaffRequired, STAFF_LIST, SWITCHED_FOR_STAFF)

    index += 1
    start_column += 1

# Print switches to day off switches sheet
print(FINAL_SWITCHES_LIST)
for i in range(0, len(FINAL_SWITCHES_LIST)):
    DAY_OFF_SWITCHES_SHEET.cell(row=i+1, column=1, value=FINAL_SWITCHES_LIST[i])

Entry.save("BookForProgramDirector.xlsx")
